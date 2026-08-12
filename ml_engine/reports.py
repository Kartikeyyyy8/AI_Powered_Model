"""
reports.py
----------
Report Generation Module for the ML Data Quality Engine.

Responsibilities
----------------
Generate an Excel audit report with the following sheets:
    - Cleaned_Data     : processed dataset (up to 10,000 rows)
    - Validation       : validation violations with count, percentage, severity
    - Business_Rules   : business rule violations with sample indices (max 10)
    - Anomalies        : ML anomaly detection model comparison
    - Quality_Scores   : dataset and column-level quality scores
    - Statistics       : descriptive statistics for numeric columns
    - Missing_Summary  : column-level missing count and missing percentage

Excel formatting applied to every sheet:
    - Bold, coloured header row
    - Auto-fit column widths
    - Freeze first row (so headers stay visible while scrolling)

Generate a JSON pipeline report including:
    - timestamp
    - pipeline_execution_summary (stage timings, total time, status)
    - validation_summary  (pass/fail stats)
    - business_rules_summary (total violations, quality score)
    - anomaly_summary (model comparison)
    - quality_scores (dataset + column scores)
    - statistics (descriptive stats + missing summary)

All report data comes from actual pipeline outputs — no mock data.

Usage
-----
    from reports import run_reports

    paths = run_reports(
        df_clean,
        violations=violations,
        scores=scores,
        val_result=val_result,
        anomaly_result=anomaly_result,
        stats_result=stats_result,
        results=pipeline_results,
        output_dir="reports",
    )
    # paths["exported"]["excel"] and paths["exported"]["json"]
"""

import json
import logging
import os
from datetime import datetime
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import LOG_FORMAT, REPORTS_DIR

import matplotlib.pyplot as plt


# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(level=logging.INFO, format=LOG_FORMAT)
logger = logging.getLogger("ml_engine.reports")

# Maximum sample indices to include in reports
_MAX_SAMPLE: int = 10

# Header style constants for Excel
_HEADER_FONT     = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL     = PatternFill(fill_type="solid", fgColor="2E4057")   # dark navy
_HEADER_ALIGN    = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Internal Helpers
# ---------------------------------------------------------------------------


def _ensure_dir(file_path: str) -> None:
    """Create the parent directory of ``file_path`` if it does not exist."""
    parent = os.path.dirname(file_path)
    if parent and not os.path.exists(parent):
        os.makedirs(parent, exist_ok=True)


def _make_timestamp() -> str:
    """Return a compact timestamp string suitable for filenames."""
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _safe_json(obj: Any) -> Any:
    """Recursively make an object JSON-serialisable."""
    if isinstance(obj, dict):
        return {k: _safe_json(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_safe_json(v) for v in obj]
    if isinstance(obj, float) and (obj != obj):  # NaN check
        return None
    if hasattr(obj, "isoformat"):
        return obj.isoformat()
    if isinstance(obj, (int, float, str, bool, type(None))):
        return obj
    return str(obj)


def _slim_violation(v: dict) -> dict:
    """
    Return a compact violation dict safe for JSON/Excel.

    Replaces large ``affected_indices`` lists with ``sample_indices`` (max 10).
    Keeps ``count``, ``percentage``, ``severity``.
    """
    slim = {
        "rule":           v.get("rule", ""),
        "column":         v.get("column", ""),
        "count":          v.get("count", 0),
        "percentage":     v.get("percentage", 0.0),
        "severity":       v.get("severity", ""),
        "sample_indices": (v.get("sample_indices") or v.get("affected_indices", []))[:_MAX_SAMPLE],
    }
    # Preserve extra diagnostic fields (e.g. example_values, unknown_values)
    for key in ("example_values", "unknown_values"):
        if key in v:
            slim[key] = v[key]
    return slim


def _format_sheet(ws, header_row: int = 1) -> None:
    """
    Apply Excel formatting to a worksheet:
    - Bold + coloured header row
    - Auto-fit column widths
    - Freeze the header row

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
    header_row : int
        Row number of the header. Default: 1.
    """
    # Bold + colour the header
    for cell in ws[header_row]:
        cell.font  = _HEADER_FONT
        cell.fill  = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN

    # Auto-fit column widths based on content (cap at 60 chars wide)
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in col_cells:
            try:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    # Freeze the header row
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


# ---------------------------------------------------------------------------
# Excel Report
# ---------------------------------------------------------------------------


def export_excel_report(
    df: pd.DataFrame,
    violations: Optional[List[dict]] = None,
    scores: Optional[Dict] = None,
    val_result: Optional[Dict] = None,
    anomaly_result: Optional[Dict] = None,
    stats_result: Optional[Dict] = None,
    output_path: str = "reports/audit_report.xlsx",
) -> str:
    """
    Export a multi-sheet, formatted Excel audit report from pipeline outputs.

    Sheets
    ------
    - **Cleaned_Data**    : processed dataframe (up to 10,000 rows)
    - **Validation**      : violations with count, percentage, severity
    - **Business_Rules**  : violations with count, %, severity, sample_indices
    - **Anomalies**       : Isolation Forest / LOF / Consensus model comparison
    - **Quality_Scores**  : dataset, column, and record-level scores
    - **Statistics**      : descriptive stats for numeric columns
    - **Missing_Summary** : column, missing count, and missing percentage

    Excel formatting applied per sheet:
    - Bold navy header row
    - Auto-fit column widths
    - Frozen header row

    Parameters
    ----------
    df : pd.DataFrame           Cleaned dataframe.
    violations : list | None    Output from ``rules.run_business_rules()``.
    scores : dict | None        Output from ``scoring.run_scoring()``.
    val_result : dict | None    Output from ``validation.run_validation()``.
    anomaly_result : dict | None Output from ``anomaly.run_ml_anomalies()``.
    stats_result : dict | None  Output from ``statistics.run_statistics()``.
    output_path : str           Destination file path.

    Returns
    -------
    str
        Absolute path of the saved Excel file.
    """
    _ensure_dir(output_path)

    try:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            wb = writer.book  # openpyxl workbook

            # ----------------------------------------------------------------
            # Sheet 1 — Cleaned_Data
            # ----------------------------------------------------------------
            df.head(10_000).to_excel(writer, sheet_name="Cleaned_Data", index=False)
            _format_sheet(wb["Cleaned_Data"])

            # ----------------------------------------------------------------
            # Sheet 2 — Validation
            # ----------------------------------------------------------------
            val_rows: List[dict] = []
            if val_result:
                for v in val_result.get("violations", []):
                    val_rows.append({
                        "Check":      v.get("check", ""),
                        "Rule":       v.get("rule", ""),
                        "Column":     v.get("column", ""),
                        "Count":      v.get("count", 0),
                        "Percentage": v.get("percentage", 0.0),
                        "Severity":   v.get("severity", ""),
                    })

            if not val_rows:
                val_rows = [{
                    "Check": "All Checks", "Rule": "No validation violations detected",
                    "Column": "N/A", "Count": 0, "Percentage": 0.0, "Severity": "Low",
                }]

            pd.DataFrame(val_rows).to_excel(writer, sheet_name="Validation", index=False)
            _format_sheet(wb["Validation"])

            # ----------------------------------------------------------------
            # Sheet 3 — Business_Rules
            # ----------------------------------------------------------------
            rule_rows: List[dict] = []
            if violations:
                for v in violations:
                    rule_rows.append({
                        "Rule":          v.get("rule", ""),
                        "Column":        v.get("column", ""),
                        "Count":         v.get("count", 0),
                        "Percentage":    v.get("percentage", 0.0),
                        "Severity":      v.get("severity", "High"),
                        "Sample_Indices": str(
                            (v.get("sample_indices") or v.get("affected_indices", []))[:_MAX_SAMPLE]
                        ),
                    })

            if not rule_rows:
                rule_rows = [{
                    "Rule": "No business rule violations detected",
                    "Column": "N/A", "Count": 0, "Percentage": 0.0,
                    "Severity": "Low", "Sample_Indices": "[]",
                }]

            pd.DataFrame(rule_rows).to_excel(writer, sheet_name="Business_Rules", index=False)
            _format_sheet(wb["Business_Rules"])

            # ----------------------------------------------------------------
            # Sheet 4 — Anomalies
            # ----------------------------------------------------------------
            anom_rows: List[dict] = []
            if anomaly_result and isinstance(anomaly_result, dict):
                features_str = ", ".join(anomaly_result.get("features_used", []))
                # Prefer model_comparison list (three-model output)
                model_comparison = anomaly_result.get("model_comparison")
                if model_comparison:
                    for mc in model_comparison:
                        anom_rows.append({
                            "Model":             mc.get("model", ""),
                            "Anomalies":         mc.get("anomalies", 0),
                            "Percentage":        mc.get("anomaly_pct", 0.0),
                            "Execution_Time_s":  mc.get("execution_time_sec", 0.0),
                            "Rows_Analysed":     mc.get("rows_analysed", 0),
                            "Features":          features_str,
                        })
                else:
                    # Fallback for older two-model output
                    anom_rows = [
                        {
                            "Model": "Isolation Forest",
                            "Anomalies": anomaly_result.get("isolation_forest_anomalies", 0),
                            "Percentage": anomaly_result.get("isolation_forest_pct", 0.0),
                            "Execution_Time_s": 0.0,
                            "Rows_Analysed": anomaly_result.get("total_rows_analysed", 0),
                            "Features": features_str,
                        },
                        {
                            "Model": "Local Outlier Factor",
                            "Anomalies": anomaly_result.get("lof_anomalies", 0),
                            "Percentage": anomaly_result.get("lof_pct", 0.0),
                            "Execution_Time_s": 0.0,
                            "Rows_Analysed": anomaly_result.get("total_rows_analysed", 0),
                            "Features": features_str,
                        },
                        {
                            "Model": "One-Class SVM",
                            "Anomalies": anomaly_result.get("one_class_svm_anomalies", 0),
                            "Percentage": anomaly_result.get("one_class_svm_pct", 0.0),
                            "Execution_Time_s": 0.0,
                            "Rows_Analysed": anomaly_result.get("total_rows_analysed", 0),
                            "Features": features_str,
                        },
                    ]
                # Always append consensus row
                anom_rows.append({
                    "Model": "Consensus (>= 2/3 models)",
                    "Anomalies": anomaly_result.get("consensus_anomalies", 0),
                    "Percentage": anomaly_result.get("consensus_pct", 0.0),
                    "Execution_Time_s": None,
                    "Rows_Analysed": anomaly_result.get("total_rows_analysed", 0),
                    "Features": features_str,
                })
            if not anom_rows:
                anom_rows = [{"Model": "N/A", "Anomalies": 0, "Percentage": 0.0,
                              "Execution_Time_s": 0.0, "Rows_Analysed": 0, "Features": "N/A"}]

            pd.DataFrame(anom_rows).to_excel(writer, sheet_name="Anomalies", index=False)
            _format_sheet(wb["Anomalies"])

            # ----------------------------------------------------------------
            # Sheet 5 — Quality_Scores
            # ----------------------------------------------------------------
            score_rows: List[dict] = []
            if scores:
                dataset_scores = scores.get("dataset", {})
                # Dataset-level scores first
                for k, v in dataset_scores.items():
                    score_rows.append({"Category": "Dataset", "Metric": k, "Value": v})
                # Column completeness
                for col, val in scores.get("columns", {}).items():
                    score_rows.append({
                        "Category": "Column Completeness",
                        "Metric": col,
                        "Value": val,
                    })
                # Record stats
                for k, v in scores.get("record_score_stats", {}).items():
                    score_rows.append({"Category": "Record Score", "Metric": k, "Value": v})

            if not score_rows:
                score_rows = [{"Category": "N/A", "Metric": "N/A", "Value": "No scoring data"}]

            pd.DataFrame(score_rows).to_excel(writer, sheet_name="Quality_Scores", index=False)
            _format_sheet(wb["Quality_Scores"])

            # ----------------------------------------------------------------
            # Sheet 6 — Statistics
            # ----------------------------------------------------------------
            stat_rows: List[dict] = []
            if stats_result:
                for col, stat in stats_result.get("descriptive_stats", {}).items():
                    row = {"Column": col}
                    row.update(stat)
                    stat_rows.append(row)

            if not stat_rows:
                stat_rows = [{"Column": "N/A", "Note": "No statistics data"}]

            pd.DataFrame(stat_rows).to_excel(writer, sheet_name="Statistics", index=False)
            _format_sheet(wb["Statistics"])

            # ----------------------------------------------------------------
            # Sheet 7 — Missing_Summary
            # ----------------------------------------------------------------
            missing_rows: List[dict] = []
            if stats_result:
                for col, minfo in stats_result.get("missing_summary", {}).items():
                    missing_rows.append({
                        "Column":             col,
                        "Missing Count":      minfo.get("null_count", 0),
                        "Missing Percentage": minfo.get("null_percentage", 0.0),
                    })

            if not missing_rows:
                missing_rows = [{"Column": "N/A", "Missing Count": 0, "Missing Percentage": 0.0}]

            pd.DataFrame(missing_rows).to_excel(writer, sheet_name="Missing_Summary", index=False)
            _format_sheet(wb["Missing_Summary"])

        abs_path = os.path.abspath(output_path)
        logger.info("Excel report saved to: %s", abs_path)
        return abs_path

    except Exception:
        logger.exception("Failed to write Excel report to '%s'.", output_path)
        raise


# ---------------------------------------------------------------------------
# JSON Report
# ---------------------------------------------------------------------------


def export_json_report(
    results: Dict[str, Any],
    scores: Optional[Dict] = None,
    val_result: Optional[Dict] = None,
    violations: Optional[List[dict]] = None,
    anomaly_result: Optional[Dict] = None,
    stats_result: Optional[Dict] = None,
    output_path: str = "reports/pipeline_report.json",
) -> str:
    """
    Export a comprehensive JSON pipeline report.

    Report sections
    ---------------
    - ``timestamp``                — ISO-format generation time
    - ``pipeline_execution_summary`` — stage timings, total time, status
    - ``validation_summary``       — pass/fail stats (from val_result)
    - ``business_rules_summary``   — total violations, quality score
    - ``anomaly_summary``          — IsoForest / LOF / Consensus counts
    - ``quality_scores``           — dataset, column, and record scores
    - ``statistics``               — descriptive stats + missing summary
    - ``validation_violations``    — compact violation list (no huge index arrays)
    - ``business_rules``           — compact violation list (sample_indices ≤ 10)

    Parameters
    ----------
    results : dict          Full pipeline result dict (from ``main.py``).
    scores : dict | None    Quality scores from ``scoring.run_scoring()``.
    val_result : dict | None Validation result from ``validation.run_validation()``.
    violations : list | None Business rule violations from ``rules.run_business_rules()``.
    anomaly_result : dict | None Anomaly detection result.
    stats_result : dict | None Statistics result.
    output_path : str       Destination file path.

    Returns
    -------
    str
        Absolute path of the saved JSON file.
    """
    _ensure_dir(output_path)

    # Build pipeline execution summary from stage timings
    stages = results.get("stages", {})
    stage_timings: Dict[str, Any] = {}
    total_time = 0.0
    for stage, info in stages.items():
        elapsed = info.get("elapsed_sec", 0.0)
        total_time += elapsed
        stage_timings[stage] = {
            "status":      "success" if info.get("error") is None else "failed",
            "elapsed_sec": elapsed,
            "error":       info.get("error"),
        }

    # Compact violations (drop large index arrays)
    compact_val_violations = [
        _slim_violation(v) for v in (val_result.get("violations", []) if val_result else [])
    ]
    compact_rule_violations = [
        _slim_violation(v) for v in (violations or [])
    ]

    # Business rules summary
    rules_quality_score = (
        scores.get("dataset", {}).get("rules_quality_score", 0.0) if scores else 0.0
    )
    business_rules_summary = {
        "total_violation_types":  len(compact_rule_violations),
        "total_affected_records": sum(v.get("count", 0) for v in (violations or [])),
        "rules_quality_score":    rules_quality_score,
        "violations":             compact_rule_violations,
    }

    # Anomaly summary: exclude only the large raw index list; keep anomaly_records
    # (already capped at 200 by anomaly.py) so the backend API can serve them.
    anomaly_summary: Dict[str, Any] = {}
    if anomaly_result and isinstance(anomaly_result, dict):
        anomaly_summary = {
            k: v
            for k, v in anomaly_result.items()
            if k not in ("consensus_indices",)  # strip only the raw index list
        }

    report = {
        "timestamp":                 datetime.now().isoformat(),
        "pipeline_execution_summary": {
            "total_stages":   len(stages),
            "total_time_sec": round(total_time, 3),
            "status":         results.get("status", "unknown"),
            "started":        results.get("pipeline_start", ""),
            "ended":          results.get("pipeline_end", ""),
            "stages":         stage_timings,
        },
        "validation_summary":        val_result.get("summary", {}) if val_result else {},
        "validation_violations":     compact_val_violations,
        "business_rules_summary":    business_rules_summary,
        "anomaly_summary":           anomaly_summary,
        "quality_scores":            scores or {},
        "statistics": {
            "descriptive_stats": stats_result.get("descriptive_stats", {}) if stats_result else {},
            "missing_summary":   stats_result.get("missing_summary", {}) if stats_result else {},
        },
    }

    try:
        with open(output_path, "w", encoding="utf-8") as fh:
            json.dump(_safe_json(report), fh, indent=2, default=str)

        abs_path = os.path.abspath(output_path)
        logger.info("JSON report saved to: %s", abs_path)
        return abs_path

    except Exception:
        logger.exception("Failed to write JSON report to '%s'.", output_path)
        raise


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------


def export_charts(df, violations=None, scores=None, output_dir="reports"):
    os.makedirs(output_dir, exist_ok=True)

    # Missing values
    missing = df.isnull().sum()

    plt.figure(figsize=(10,5))
    missing.plot(kind="bar")
    plt.title("Missing Values by Column")
    plt.ylabel("Count")
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, "missing_values.png"))
    plt.close()

    # Severity distribution
    if violations:
        violations_df = pd.DataFrame(violations)

        if "severity" in violations_df.columns:
            plt.figure(figsize=(6,4))
            violations_df["severity"].value_counts().plot(kind="bar")
            plt.title("Business Rule Violations by Severity")
            plt.tight_layout()
            plt.savefig(os.path.join(output_dir, "severity_distribution.png"))
            plt.close()

    # Dataset score
    if scores:
        dataset_score = scores["dataset"]["dataset_score"]

        plt.figure(figsize=(5,4))
        plt.bar(["Dataset"], [dataset_score])
        plt.ylim(0,100)
        plt.ylabel("Score")
        plt.title("Overall Dataset Quality Score")
        plt.tight_layout()
        plt.savefig(os.path.join(output_dir, "dataset_score.png"))
        plt.close()

    return {
    "missing_values": os.path.join(output_dir, "missing_values.png"),
    "severity_distribution": os.path.join(output_dir, "severity_distribution.png"),
    "dataset_score": os.path.join(output_dir, "dataset_score.png"),
}


def run_reports(
    df: pd.DataFrame,
    violations: Optional[List[dict]] = None,
    scores: Optional[Dict] = None,
    val_result: Optional[Dict] = None,
    anomaly_result: Optional[Dict] = None,
    stats_result: Optional[Dict] = None,
    results: Optional[Dict] = None,
    output_dir: str = REPORTS_DIR,
) -> Dict[str, str]:
    """
    Generate both Excel and JSON reports and return their file paths.

    Parameters
    ----------
    df : pd.DataFrame           Cleaned dataframe.
    violations : list | None    Business rule violations.
    scores : dict | None        Quality scores.
    val_result : dict | None    Validation result.
    anomaly_result : dict | None Anomaly detection result.
    stats_result : dict | None  Statistics result.
    results : dict | None       Full pipeline results dict for JSON timing section.
    output_dir : str            Directory where reports are saved.

    Returns
    -------
    dict
        ``{"status": "success", "exported": {"excel": "...", "json": "..."}}``
    """
    ts = _make_timestamp()
    os.makedirs(output_dir, exist_ok=True)

    excel_path = os.path.join(output_dir, f"audit_report_{ts}.xlsx")
    json_path  = os.path.join(output_dir, f"pipeline_report_{ts}.json")

    excel_abs = export_excel_report(
        df,
        violations=violations,
        scores=scores,
        val_result=val_result,
        anomaly_result=anomaly_result,
        stats_result=stats_result,
        output_path=excel_path,
    )

    json_abs = export_json_report(
        results=results or {},
        scores=scores,
        val_result=val_result,
        violations=violations,
        anomaly_result=anomaly_result,
        stats_result=stats_result,
        output_path=json_path,
    )

    chart_paths = export_charts(
    df=df,
    violations=violations,
    scores=scores,
    output_dir=output_dir,
    )

    report_paths = {
    "status": "success",
    "exported": {
        "excel": excel_abs,
        "json": json_abs,
        "charts": chart_paths,
    },
}
    logger.info("Charts exported to: %s", output_dir)
    logger.info("Reports generated: %s", list(report_paths["exported"].values()))
    return report_paths
    



